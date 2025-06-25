#!/usr/bin/env python3
"""
Script simple para crear el archivo de resultados usando solo openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Configuración de columnas según config.py
EVALUATIONS_COLUMNS = {
    "evaluation_id": "A",
    "cedula": "B",
    "nombre": "C", 
    "cargo": "D",
    "campo": "E",
    "procedure_codigo": "F",
    "procedure_nombre": "G",
    "total_questions": "H",
    "correct_answers": "I",
    "score_percentage": "J",
    "aprobo": "K",
    "completed_at": "L"
}

ANSWERS_COLUMNS = {
    "evaluation_id": "A",
    "question_id": "B",
    "question_text": "C",
    "selected_option": "D",
    "selected_text": "E",
    "correct_option": "F",
    "correct_text": "G",
    "is_correct": "H",
    "display_order_question": "I",
    "display_order_a": "J",
    "display_order_b": "K",
    "display_order_c": "L",
    "display_order_d": "M"
}

APPLIED_KNOWLEDGE_COLUMNS = {
    "evaluation_id": "A",
    "describio_procedimiento": "B",
    "identifico_riesgos": "C",
    "identifico_epp": "D",
    "describio_incidentes": "E"
}

FEEDBACK_COLUMNS = {
    "evaluation_id": "A",
    "hizo_sugerencia": "B",
    "cual_sugerencia": "C",
    "aprobo": "D",
    "requiere_entrenamiento": "E"
}

def get_col_index(column_letter):
    """Convertir letra de columna a índice (A=0, B=1, etc.)"""
    return ord(column_letter.upper()) - ord('A')

def create_results_file():
    """Crear archivo de resultados con headers"""
    
    # Ruta del archivo
    data_dir = Path(__file__).parent / "data"
    results_file = data_dir / "resultados_evaluaciones.xlsx"
    
    print(f"📝 Creando archivo: {results_file}")
    
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Crear hojas con headers
    sheets_config = [
        ("Evaluaciones", EVALUATIONS_COLUMNS),
        ("Respuestas", ANSWERS_COLUMNS),
        ("Conocimiento_Aplicado", APPLIED_KNOWLEDGE_COLUMNS),
        ("Feedback", FEEDBACK_COLUMNS)
    ]
    
    for sheet_name, columns_config in sheets_config:
        print(f"  📋 Creando hoja: {sheet_name}")
        ws = wb.create_sheet(title=sheet_name)
        
        # Escribir headers
        for field_name, column_letter in columns_config.items():
            col_index = get_col_index(column_letter)
            header_text = field_name.replace("_", " ").title()
            ws.cell(row=1, column=col_index + 1, value=header_text)
            
            # Estilo para headers
            cell = ws.cell(row=1, column=col_index + 1)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    wb.save(results_file)
    wb.close()
    print(f"✅ Archivo creado exitosamente: {results_file}")

if __name__ == "__main__":
    create_results_file()