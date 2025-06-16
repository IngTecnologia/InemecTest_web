"""
Módulo de sincronización con Excel para integrar preguntas generadas
backend/src/admin/excel_sync.py
"""

import os
import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from ..excel_handler import ExcelHandler
from ..config import (
    get_data_file_path,
    DATA_SHEETS,
    QUESTIONS_COLUMNS,
    ensure_data_directory
)
from .models import QuestionBatch, QuestionInProcess

class ExcelSyncManager:
    """
    Gestiona la sincronización de preguntas generadas con el archivo Excel principal
    """
    
    def __init__(self):
        self.excel_handler = ExcelHandler()
        self.data_file = get_data_file_path()
        ensure_data_directory()
        
        print(f"📊 ExcelSyncManager inicializado:")
        print(f"   - Archivo Excel: {self.data_file}")
    
    async def sync_batch_to_excel(self, batch: QuestionBatch) -> Dict[str, Any]:
        """
        Sincronizar un lote de preguntas al archivo Excel
        """
        try:
            print(f"📊 Sincronizando lote {batch.batch_id} al Excel...")
            
            # 1. Verificar/crear archivo Excel si no existe
            await self._ensure_excel_file_exists()
            
            # 2. Agregar/actualizar procedimiento en hoja de Procedimientos
            await self._sync_procedure_to_excel(batch)
            
            # 3. Agregar preguntas a hoja de Preguntas
            questions_added = await self._sync_questions_to_excel(batch)
            
            # 4. Guardar backup del archivo JSON temporal también
            await self._save_generated_questions_json(batch)
            
            result = {
                "success": True,
                "batch_id": batch.batch_id,
                "procedure_codigo": batch.procedure_codigo,
                "questions_synced": questions_added,
                "excel_file": str(self.data_file),
                "timestamp": datetime.now().isoformat()
            }
            
            print(f"✅ Sincronización completada: {questions_added} preguntas añadidas al Excel")
            return result
            
        except Exception as e:
            print(f"❌ Error sincronizando con Excel: {e}")
            return {
                "success": False,
                "error": str(e),
                "batch_id": batch.batch_id
            }
    
    async def _ensure_excel_file_exists(self):
        """
        Verificar que el archivo Excel existe, crearlo si no
        """
        if not self.data_file.exists():
            print(f"📄 Creando archivo Excel: {self.data_file}")
            await self._create_initial_excel_file()
        else:
            print(f"✅ Archivo Excel encontrado: {self.data_file}")
    
    async def _create_initial_excel_file(self):
        """
        Crear archivo Excel inicial con las hojas necesarias
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Crear hoja de Procedimientos
        ws_proc = wb.create_sheet(title=DATA_SHEETS["procedures"]["name"])
        
        # Headers para Procedimientos
        headers_proc = ["Código", "Nombre", "Alcance", "Objetivo"]
        for i, header in enumerate(headers_proc, 1):
            cell = ws_proc.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Crear hoja de Preguntas
        ws_quest = wb.create_sheet(title=DATA_SHEETS["questions"]["name"])
        
        # Headers para Preguntas
        headers_quest = ["Código Procedimiento", "Pregunta", "Opción A", "Opción B", "Opción C", "Opción D"]
        for i, header in enumerate(headers_quest, 1):
            cell = ws_quest.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Guardar archivo
        wb.save(self.data_file)
        wb.close()
        
        print(f"✅ Archivo Excel creado: {self.data_file}")
    
    async def _sync_procedure_to_excel(self, batch: QuestionBatch):
        """
        Agregar o actualizar procedimiento en la hoja de Procedimientos
        """
        try:
            # Leer hoja de procedimientos existente
            try:
                df_proc = pd.read_excel(self.data_file, sheet_name=DATA_SHEETS["procedures"]["name"])
            except Exception:
                # Si no existe la hoja, crear DataFrame vacío
                df_proc = pd.DataFrame(columns=["Código", "Nombre", "Alcance", "Objetivo"])
            
            # Verificar si el procedimiento ya existe
            if not df_proc.empty and batch.procedure_codigo in df_proc["Código"].values:
                print(f"   📝 Procedimiento {batch.procedure_codigo} ya existe en Excel")
                return
            
            # Agregar nuevo procedimiento
            new_row = {
                "Código": batch.procedure_codigo,
                "Nombre": batch.procedure_name or f"Procedimiento {batch.procedure_codigo}",
                "Alcance": f"Procedimiento técnico versión {batch.procedure_version}",
                "Objetivo": f"Ejecutar procedimiento {batch.procedure_codigo} según especificaciones técnicas"
            }
            
            df_proc = pd.concat([df_proc, pd.DataFrame([new_row])], ignore_index=True)
            
            # Guardar de vuelta al Excel
            with pd.ExcelWriter(self.data_file, mode='a', if_sheet_exists='replace') as writer:
                df_proc.to_excel(writer, sheet_name=DATA_SHEETS["procedures"]["name"], index=False)
            
            print(f"   ✅ Procedimiento {batch.procedure_codigo} agregado al Excel")
            
        except Exception as e:
            print(f"   ⚠️ Error sincronizando procedimiento: {e}")
    
    async def _sync_questions_to_excel(self, batch: QuestionBatch) -> int:
        """
        Agregar preguntas a la hoja de Preguntas
        """
        try:
            # Leer hoja de preguntas existente
            try:
                df_quest = pd.read_excel(self.data_file, sheet_name=DATA_SHEETS["questions"]["name"])
            except Exception:
                # Si no existe la hoja, crear DataFrame vacío
                df_quest = pd.DataFrame(columns=["Código Procedimiento", "Pregunta", "Opción A", "Opción B", "Opción C", "Opción D"])
            
            # Preparar nuevas preguntas
            new_questions = []
            for question in batch.questions:
                if question.status.value in ["completed", "needs_correction"]:  # Solo preguntas válidas
                    new_row = {
                        "Código Procedimiento": question.procedure_codigo,
                        "Pregunta": question.pregunta,
                        "Opción A": question.opciones[0],  # Opción correcta siempre en A
                        "Opción B": question.opciones[1],
                        "Opción C": question.opciones[2],
                        "Opción D": question.opciones[3]
                    }
                    new_questions.append(new_row)
            
            if not new_questions:
                print(f"   ⚠️ No hay preguntas válidas para sincronizar en lote {batch.batch_id}")
                return 0
            
            # Agregar nuevas preguntas
            df_new = pd.DataFrame(new_questions)
            df_quest = pd.concat([df_quest, df_new], ignore_index=True)
            
            # Guardar de vuelta al Excel
            with pd.ExcelWriter(self.data_file, mode='a', if_sheet_exists='replace') as writer:
                df_quest.to_excel(writer, sheet_name=DATA_SHEETS["questions"]["name"], index=False)
            
            print(f"   ✅ {len(new_questions)} preguntas agregadas al Excel")
            return len(new_questions)
            
        except Exception as e:
            print(f"   ❌ Error sincronizando preguntas: {e}")
            return 0
    
    async def _save_generated_questions_json(self, batch: QuestionBatch):
        """
        Guardar también en formato JSON compatible con el sistema anterior
        """
        try:
            # Preparar archivo de preguntas generadas
            generated_file = Path("backend/data/generated_questions.json")
            
            # Cargar existente o crear nuevo
            if generated_file.exists():
                with open(generated_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            else:
                existing_data = []
            
            # Agregar nuevas preguntas en formato compatible
            for question in batch.questions:
                if question.status.value in ["completed", "needs_correction"]:
                    question_data = {
                        "codigo_procedimiento": question.procedure_codigo,
                        "version_proc": int(question.procedure_version),
                        "version_preg": question.version_preg,
                        "prompt": question.prompt,
                        "puntaje_ia": question.puntaje_ia,
                        "puntaje_e1": 0,
                        "puntaje_e2": 0,
                        "puntaje_e3": 0,
                        "puntaje_e4": 0,
                        "comentario_e1": "",
                        "comentario_e2": "",
                        "comentario_e3": "",
                        "comentario_e4": "",
                        "pregunta": question.pregunta,
                        "opciones": question.opciones,
                        "historial_revision": question.historial_revision,
                        "batch_id": batch.batch_id,
                        "generated_at": question.created_at,
                        "validation_score": self._calculate_question_validation_score(question)
                    }
                    existing_data.append(question_data)
            
            # Guardar archivo actualizado
            with open(generated_file, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, indent=2, ensure_ascii=False)
            
            print(f"   📄 Preguntas guardadas en: {generated_file}")
            
        except Exception as e:
            print(f"   ⚠️ Error guardando JSON: {e}")
    
    def _calculate_question_validation_score(self, question: QuestionInProcess) -> float:
        """
        Calcular score de validación de una pregunta
        """
        if not question.validations:
            return 0.0
        
        total_score = 0
        total_weight = 0
        
        for validation in question.validations:
            # Usar peso por defecto de 1.0 si no está en config
            weight = 1.0
            total_score += validation.score * weight
            total_weight += weight
        
        return total_score / total_weight if total_weight > 0 else 0.0

# Funciones de utilidad
def create_excel_sync_manager() -> ExcelSyncManager:
    """Crear instancia del sync manager"""
    return ExcelSyncManager()

async def sync_batch_to_excel(batch: QuestionBatch) -> Dict[str, Any]:
    """Función de conveniencia para sincronizar un lote"""
    sync_manager = create_excel_sync_manager()
    return await sync_manager.sync_batch_to_excel(batch)