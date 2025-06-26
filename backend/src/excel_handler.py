"""
Manejador de archivos Excel para InemecTest
Versión completa y funcional basada en Excel
"""

import pandas as pd
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .config import (
    get_data_file_path, 
    get_results_file_path,
    DATA_SHEETS,
    RESULTS_SHEETS,
    PROCEDURES_COLUMNS,
    QUESTIONS_COLUMNS,
    EVALUATIONS_COLUMNS,
    ANSWERS_COLUMNS,
    APPLIED_KNOWLEDGE_COLUMNS,
    FEEDBACK_COLUMNS,
    VALID_CAMPOS,
    VALID_OPTIONS,
    VALID_SI_NO,
    ensure_data_directory
)

class ExcelHandler:
    """Clase para manejar todas las operaciones con Excel"""
    
    def __init__(self):
        ensure_data_directory()
        self.data_file = get_data_file_path()
        self.results_file = get_results_file_path()
        print(f"📁 Excel Handler inicializado:")
        print(f"   - Archivo de datos: {self.data_file}")
        print(f"   - Archivo de resultados: {self.results_file}")
    
    # =================================================================
    # LECTURA DE DATOS (Procedimientos y Preguntas)
    # =================================================================
    
    async def get_all_procedures(self) -> List[Dict[str, Any]]:
        """Obtener todos los procedimientos desde Excel"""
        try:
            print(f"🔍 [DEBUG] Buscando archivo de datos en: {self.data_file}")
            print(f"🔍 [DEBUG] Archivo existe: {self.data_file.exists()}")
            
            if not self.data_file.exists():
                print(f"⚠️ Archivo de datos no encontrado: {self.data_file}")
                return []
            
            # Leer hoja de procedimientos
            df = pd.read_excel(self.data_file, sheet_name=DATA_SHEETS["procedures"]["name"])
            
            procedures = []
            for index, row in df.iterrows():
                # Saltar filas completamente vacías
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == "":
                    continue
                
                try:
                    # Obtener campos básicos
                    procedure = {
                        "codigo": str(row.iloc[self._get_col_index(PROCEDURES_COLUMNS["codigo"])]).strip(),
                        "nombre": str(row.iloc[self._get_col_index(PROCEDURES_COLUMNS["nombre"])]).strip(),
                        "alcance": str(row.iloc[self._get_col_index(PROCEDURES_COLUMNS["alcance"])]).strip(),
                        "objetivo": str(row.iloc[self._get_col_index(PROCEDURES_COLUMNS["objetivo"])]).strip()
                    }
                    
                    # Obtener campos adicionales para filtros
                    disciplina_col_index = self._get_col_index(PROCEDURES_COLUMNS["disciplina"])  # G = 6
                    campo_col_index = self._get_col_index(PROCEDURES_COLUMNS["campo"])  # L = 11
                    
                    # Verificar que las columnas existan en la fila
                    disciplina_raw = ""
                    campo_raw = ""
                    
                    if disciplina_col_index < len(row):
                        disciplina_raw = str(row.iloc[disciplina_col_index]).strip()
                    
                    if campo_col_index < len(row):
                        campo_raw = str(row.iloc[campo_col_index]).strip()
                    
                    # Limpiar valores NaN y vacíos
                    disciplina = disciplina_raw if disciplina_raw != "nan" and disciplina_raw != "" else None
                    campo = campo_raw if campo_raw != "nan" and campo_raw != "" else None
                    
                    # Agregar datos completos para filtros del frontend
                    procedure["datos_completos"] = {
                        "disciplina": disciplina,
                        "campo": campo
                    }
                    
                    # Validar que el código no esté vacío
                    if procedure["codigo"] and procedure["codigo"] != "nan":
                        procedures.append(procedure)
                        
                except Exception as e:
                    print(f"⚠️ Error procesando fila {index + 2}: {e}")
                    continue
            
            print(f"✅ Cargados {len(procedures)} procedimientos")
            return procedures
            
        except Exception as e:
            print(f"❌ Error leyendo procedimientos: {e}")
            return []
    
    async def get_procedure_by_code(self, codigo: str) -> Optional[Dict[str, Any]]:
        """Obtener un procedimiento específico por código"""
        procedures = await self.get_all_procedures()
        
        for proc in procedures:
            if proc["codigo"].upper() == codigo.upper():
                return proc
        
        return None
    
    async def get_questions_by_procedure(self, procedure_codigo: str) -> List[Dict[str, Any]]:
        """Obtener preguntas de un procedimiento específico"""
        try:
            if not self.data_file.exists():
                print(f"⚠️ Archivo de datos no encontrado: {self.data_file}")
                return []
            
            # Leer hoja de preguntas
            df = pd.read_excel(self.data_file, sheet_name=DATA_SHEETS["questions"]["name"])
            
            # DEBUG: Información detallada
            print(f"🔍 DEBUG - Buscando preguntas para: '{procedure_codigo}'")
            print(f"🔍 DEBUG - Columnas Excel: {list(df.columns)}")
            print(f"🔍 DEBUG - Total filas: {len(df)}")
            print(f"🔍 DEBUG - Configuración columna procedure_codigo: {QUESTIONS_COLUMNS['procedure_codigo']}")
            print(f"🔍 DEBUG - Índice columna: {self._get_col_index(QUESTIONS_COLUMNS['procedure_codigo'])}")
            
            # Mostrar primeras 3 filas de la columna procedure_codigo
            if not df.empty:
                proc_col_index = self._get_col_index(QUESTIONS_COLUMNS["procedure_codigo"])
                print(f"🔍 DEBUG - Primeros valores columna A:")
                for i in range(min(3, len(df))):
                    val = str(df.iloc[i, proc_col_index]).strip()
                    print(f"   Fila {i+2}: '{val}' (tipo: {type(df.iloc[i, proc_col_index])})")
            
            questions = []
            question_id = 1
            
            for index, row in df.iterrows():
                # Saltar filas vacías
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == "":
                    print(f"🔍 DEBUG - Saltando fila vacía {index + 2}")
                    continue
                
                try:
                    row_procedure_codigo = str(row.iloc[self._get_col_index(QUESTIONS_COLUMNS["procedure_codigo"])]).strip()
                    print(f"🔍 DEBUG - Fila {index + 2}: Comparando '{row_procedure_codigo}' vs '{procedure_codigo}'")
                    
                    if row_procedure_codigo.upper() == procedure_codigo.upper():
                        print(f"🔍 DEBUG - ¡MATCH! Procesando pregunta de fila {index + 2}")
                        
                        question = {
                            "id": question_id,
                            "procedure_codigo": row_procedure_codigo,
                            "question_text": str(row.iloc[self._get_col_index(QUESTIONS_COLUMNS["question_text"])]).strip(),
                            "option_a": str(row.iloc[self._get_col_index(QUESTIONS_COLUMNS["option_a"])]).strip(),
                            "option_b": str(row.iloc[self._get_col_index(QUESTIONS_COLUMNS["option_b"])]).strip(),
                            "option_c": str(row.iloc[self._get_col_index(QUESTIONS_COLUMNS["option_c"])]).strip(),
                            "option_d": str(row.iloc[self._get_col_index(QUESTIONS_COLUMNS["option_d"])]).strip(),
                            "correct_answer": "A"  # ← SIEMPRE A, ya que Option_A es la correcta
                        }
                        
                        print(f"🔍 DEBUG - Pregunta creada: {question}")
                        
                        # Validar que la pregunta esté completa
                        if (question["question_text"] and question["question_text"] != "nan"):
                            questions.append(question)
                            question_id += 1
                            print(f"🔍 DEBUG - Pregunta añadida exitosamente")
                        else:
                            print(f"🔍 DEBUG - Pregunta rechazada: question_text inválido")
                    else:
                        print(f"🔍 DEBUG - No match: '{row_procedure_codigo}' != '{procedure_codigo}'")
                            
                except Exception as e:
                    print(f"⚠️ Error procesando pregunta en fila {index + 2}: {e}")
                    continue
            
            print(f"✅ Cargadas {len(questions)} preguntas para {procedure_codigo}")
            return questions
            
        except Exception as e:
            print(f"❌ Error leyendo preguntas para {procedure_codigo}: {e}")
            return []
    
    # =================================================================
    # ESCRITURA DE RESULTADOS
    # =================================================================
    
    async def save_evaluation_result(self, evaluation_data: Dict[str, Any]) -> str:
        """Guardar resultado completo de evaluación en Excel"""
        try:
            # Usar cédula como identificador principal
            cedula = evaluation_data["user_data"]["cedula"]
            # Generar ID único para esta sesión específica de la cédula
            evaluation_id = f"{cedula}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Preparar todas las hojas de datos
            evaluation_row = await self._prepare_evaluation_row(evaluation_id, evaluation_data)
            answers_rows = await self._prepare_answers_rows(evaluation_id, evaluation_data)
            applied_row = await self._prepare_applied_knowledge_row(evaluation_id, evaluation_data)
            feedback_row = await self._prepare_feedback_row(evaluation_id, evaluation_data)
            
            # Escribir a Excel
            await self._write_to_results_excel(
                evaluation_row=evaluation_row,
                answers_rows=answers_rows,
                applied_row=applied_row,
                feedback_row=feedback_row
            )
            
            print(f"✅ Evaluación guardada para cédula: {cedula}, ID sesión: {evaluation_id}")
            return evaluation_id
            
        except Exception as e:
            print(f"❌ Error guardando evaluación: {e}")
            raise e
    
    async def _prepare_evaluation_row(self, evaluation_id: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """Preparar fila principal de evaluación"""
        now = datetime.now()
        score_data = data.get("score_data", {})
        
        return {
            "evaluation_id": evaluation_id,
            "cedula": data["user_data"]["cedula"],
            "nombre": data["user_data"]["nombre"],
            "cargo": data["user_data"]["cargo"],
            "campo": data["user_data"]["campo"],
            "procedure_codigo": data["procedure_codigo"],
            "procedure_nombre": data.get("procedure_nombre", ""),
            "total_questions": score_data.get("total_questions", 0),
            "correct_answers": score_data.get("correct_answers", 0),
            "score_percentage": score_data.get("score_percentage", 0),
            "aprobo": data["feedback"]["aprobo"],
            "started_at": now.strftime("%Y-%m-%d %H:%M:%S"),
            "completed_at": now.strftime("%Y-%m-%d %H:%M:%S")
        }
    
    async def _prepare_answers_rows(self, evaluation_id: str, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Preparar filas de respuestas detalladas con orden de visualización completo"""
        rows = []
        
        for i, answer in enumerate(data["knowledge_answers"], 1):
            display_order = answer.get("display_order", {})
            
            row = {
                "evaluation_id": evaluation_id,
                "question_id": answer.get("question_id", i),
                "question_text": display_order.get("question_text", answer.get("question_text", "")),
                "option_a_text": display_order.get("option_a_text", ""),
                "option_b_text": display_order.get("option_b_text", ""),
                "option_c_text": display_order.get("option_c_text", ""),
                "option_d_text": display_order.get("option_d_text", ""),
                "selected_option": answer["selected_option"],
                "selected_text": answer.get("selected_text", ""),
                "correct_option": answer.get("correct_option", ""),
                "correct_text": answer.get("correct_text", ""),
                "correct_option_displayed": answer.get("correct_option_displayed", ""),
                "is_correct": "Sí" if answer.get("is_correct", False) else "No"
            }
            rows.append(row)
        
        return rows
    
    async def _prepare_applied_knowledge_row(self, evaluation_id: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """Preparar fila de conocimiento aplicado"""
        applied = data["applied_knowledge"]
        
        return {
            "evaluation_id": evaluation_id,
            "describio_procedimiento": "Sí" if applied["describio_procedimiento"] else "No",
            "identifico_riesgos": "Sí" if applied["identifico_riesgos"] else "No",
            "identifico_epp": "Sí" if applied["identifico_epp"] else "No",
            "describio_incidentes": "Sí" if applied["describio_incidentes"] else "No"
        }
    
    async def _prepare_feedback_row(self, evaluation_id: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """Preparar fila de feedback"""
        feedback = data["feedback"]
        
        return {
            "evaluation_id": evaluation_id,
            "hizo_sugerencia": feedback["hizo_sugerencia"],
            "cual_sugerencia": feedback.get("cual_sugerencia", ""),
            "requiere_entrenamiento": feedback.get("requiere_entrenamiento", "")
        }
    
    async def _write_to_results_excel(self, **data_rows):
        """Escribir todos los datos al archivo de resultados Excel"""
        try:
            # Crear archivo si no existe
            if not self.results_file.exists():
                await self._create_results_file()
            
            # Cargar workbook existente
            wb = load_workbook(self.results_file)
            
            # Escribir en cada hoja
            await self._append_to_sheet(wb, RESULTS_SHEETS["evaluations"]["name"], 
                                      data_rows["evaluation_row"], EVALUATIONS_COLUMNS)
            
            for answer_row in data_rows["answers_rows"]:
                await self._append_to_sheet(wb, RESULTS_SHEETS["answers"]["name"], 
                                          answer_row, ANSWERS_COLUMNS)
            
            await self._append_to_sheet(wb, RESULTS_SHEETS["applied_knowledge"]["name"], 
                                      data_rows["applied_row"], APPLIED_KNOWLEDGE_COLUMNS)
            
            await self._append_to_sheet(wb, RESULTS_SHEETS["feedback"]["name"], 
                                      data_rows["feedback_row"], FEEDBACK_COLUMNS)
            
            # Guardar archivo
            wb.save(self.results_file)
            wb.close()
            
        except Exception as e:
            print(f"❌ Error escribiendo resultados: {e}")
            raise e
    
    async def _create_results_file(self):
        """Crear archivo de resultados con headers"""
        wb = Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Crear hojas con headers
        sheets_config = [
            (RESULTS_SHEETS["evaluations"]["name"], EVALUATIONS_COLUMNS),
            (RESULTS_SHEETS["answers"]["name"], ANSWERS_COLUMNS),
            (RESULTS_SHEETS["applied_knowledge"]["name"], APPLIED_KNOWLEDGE_COLUMNS),
            (RESULTS_SHEETS["feedback"]["name"], FEEDBACK_COLUMNS)
        ]
        
        for sheet_name, columns_config in sheets_config:
            ws = wb.create_sheet(title=sheet_name)
            
            # Escribir headers
            for field_name, column_letter in columns_config.items():
                col_index = self._get_col_index(column_letter)
                header_text = field_name.replace("_", " ").title()
                ws.cell(row=1, column=col_index + 1, value=header_text)
                
                # Estilo para headers
                cell = ws.cell(row=1, column=col_index + 1)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        wb.save(self.results_file)
        wb.close()
        print(f"✅ Archivo de resultados creado: {self.results_file}")
    
    async def _append_to_sheet(self, wb, sheet_name: str, data: Dict[str, Any], columns_config: Dict[str, str]):
        """Agregar fila de datos a una hoja específica"""
        ws = wb[sheet_name]
        
        # Encontrar próxima fila vacía
        next_row = ws.max_row + 1
        
        # Escribir datos en las columnas correspondientes
        for field_name, column_letter in columns_config.items():
            if field_name in data:
                col_index = self._get_col_index(column_letter)
                ws.cell(row=next_row, column=col_index + 1, value=data[field_name])
    
    # =================================================================
    # FUNCIONES DE CONSULTA DE RESULTADOS
    # =================================================================
    
    async def get_evaluation_results(self, evaluation_id: str) -> Optional[Dict[str, Any]]:
        """Obtener resultados completos de una evaluación"""
        try:
            if not self.results_file.exists():
                return None
            
            # Leer todas las hojas
            evaluation_df = pd.read_excel(self.results_file, sheet_name=RESULTS_SHEETS["evaluations"]["name"])
            answers_df = pd.read_excel(self.results_file, sheet_name=RESULTS_SHEETS["answers"]["name"])
            applied_df = pd.read_excel(self.results_file, sheet_name=RESULTS_SHEETS["applied_knowledge"]["name"])
            feedback_df = pd.read_excel(self.results_file, sheet_name=RESULTS_SHEETS["feedback"]["name"])
            
            # Filtrar por evaluation_id
            evaluation_row = evaluation_df[evaluation_df['Evaluation Id'] == evaluation_id]
            if evaluation_row.empty:
                return None
            
            answers_rows = answers_df[answers_df['Evaluation Id'] == evaluation_id]
            applied_row = applied_df[applied_df['Evaluation Id'] == evaluation_id]
            feedback_row = feedback_df[feedback_df['Evaluation Id'] == evaluation_id]
            
            # Construir resultado
            result = {
                "evaluation": evaluation_row.iloc[0].to_dict() if not evaluation_row.empty else {},
                "answers": [row.to_dict() for _, row in answers_rows.iterrows()],
                "applied": applied_row.iloc[0].to_dict() if not applied_row.empty else {},
                "feedback": feedback_row.iloc[0].to_dict() if not feedback_row.empty else {}
            }
            
            return result
            
        except Exception as e:
            print(f"❌ Error obteniendo resultados para {evaluation_id}: {e}")
            return None
    
    async def get_all_evaluations(self) -> List[Dict[str, Any]]:
        """Obtener lista de todas las evaluaciones"""
        try:
            if not self.results_file.exists():
                return []
            
            df = pd.read_excel(self.results_file, sheet_name=RESULTS_SHEETS["evaluations"]["name"])
            
            # Mapear nombres de columnas del Excel a nombres esperados por el código
            column_mapping = {
                'Evaluation Id': 'evaluation_id',
                'Cedula': 'cedula',
                'Nombre': 'nombre',
                'Cargo': 'cargo',
                'Campo': 'campo',
                'Procedure Codigo': 'procedure_codigo',
                'Procedure Nombre': 'procedure_nombre',
                'Total Questions': 'total_questions',
                'Correct Answers': 'correct_answers',
                'Score Percentage': 'score_percentage',
                'Aprobo': 'aprobo',
                'Started At': 'started_at',
                'Completed At': 'completed_at'
            }
            
            # Renombrar columnas
            df = df.rename(columns=column_mapping)
            
            evaluations = [row.to_dict() for _, row in df.iterrows()]
            
            # Sanitizar datos para evitar objetos AdminResponse embebidos
            sanitized_evaluations = []
            for evaluation in evaluations:
                sanitized_eval = self._sanitize_evaluation_data(evaluation)
                if sanitized_eval:  # Solo agregar si la sanitización fue exitosa
                    sanitized_evaluations.append(sanitized_eval)
            
            return sanitized_evaluations
            
        except Exception as e:
            print(f"❌ Error obteniendo evaluaciones: {e}")
            return []
    
    async def get_procedure_statistics(self) -> List[Dict[str, Any]]:
        """Obtener estadísticas por procedimiento"""
        try:
            if not self.results_file.exists():
                return []
            
            df = pd.read_excel(self.results_file, sheet_name=RESULTS_SHEETS["evaluations"]["name"])
            
            # Agrupar por procedimiento
            stats = []
            if not df.empty:
                for codigo in df['Procedure Codigo'].unique():
                    proc_data = df[df['Procedure Codigo'] == codigo]
                    
                    stat = {
                        "procedure_codigo": codigo,
                        "procedure_name": proc_data['Procedure Nombre'].iloc[0] if not proc_data.empty else "",
                        "total_evaluations": len(proc_data),
                        "average_score": round(proc_data['Score Percentage'].mean(), 2),
                        "approval_rate": round((proc_data['Aprobo'].str.contains('si', case=False, na=False)).sum() / len(proc_data) * 100, 2)
                    }
                    stats.append(stat)
            
            return sorted(stats, key=lambda x: x['total_evaluations'], reverse=True)
            
        except Exception as e:
            print(f"❌ Error obteniendo estadísticas: {e}")
            return []

    # =================================================================
    # MÉTODOS PARA GESTIÓN DE EVALUACIONES (ADMIN)
    # =================================================================
    
    async def get_evaluation_by_id(self, evaluation_id: str) -> Optional[Dict[str, Any]]:
        """Obtener datos principales de una evaluación por ID"""
        try:
            if not self.results_file.exists():
                return None
            
            evaluations = await self.get_all_evaluations()
            for evaluation in evaluations:
                if evaluation.get("evaluation_id") == evaluation_id:
                    return evaluation
            
            return None
            
        except Exception as e:
            print(f"❌ Error obteniendo evaluación por ID: {e}")
            return None
    
    async def get_evaluation_answers(self, evaluation_id: str) -> List[Dict[str, Any]]:
        """Obtener respuestas detalladas de una evaluación"""
        try:
            if not self.results_file.exists():
                return []
            
            wb = load_workbook(self.results_file, data_only=True)
            
            if RESULTS_SHEETS["answers"]["name"] not in wb.sheetnames:
                return []
            
            ws = wb[RESULTS_SHEETS["answers"]["name"]]
            answers = []
            
            # Leer datos de respuestas
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= len(ANSWERS_COLUMNS) and row[0] == evaluation_id:
                    answer_data = {}
                    for col_name, col_letter in ANSWERS_COLUMNS.items():
                        col_index = self._get_col_index(col_letter)
                        if col_index < len(row):
                            answer_data[col_name] = row[col_index]
                    answers.append(answer_data)
            
            wb.close()
            return sorted(answers, key=lambda x: x.get("question_id", 0))
            
        except Exception as e:
            print(f"❌ Error obteniendo respuestas de evaluación: {e}")
            return []
    
    async def get_evaluation_applied_knowledge(self, evaluation_id: str) -> Optional[Dict[str, Any]]:
        """Obtener datos de conocimiento aplicado de una evaluación"""
        try:
            if not self.results_file.exists():
                return None
            
            wb = load_workbook(self.results_file, data_only=True)
            
            if RESULTS_SHEETS["applied_knowledge"]["name"] not in wb.sheetnames:
                return None
            
            ws = wb[RESULTS_SHEETS["applied_knowledge"]["name"]]
            
            # Buscar fila con el evaluation_id
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= len(APPLIED_KNOWLEDGE_COLUMNS) and row[0] == evaluation_id:
                    applied_data = {}
                    for col_name, col_letter in APPLIED_KNOWLEDGE_COLUMNS.items():
                        col_index = self._get_col_index(col_letter)
                        if col_index < len(row):
                            applied_data[col_name] = row[col_index]
                    wb.close()
                    return applied_data
            
            wb.close()
            return None
            
        except Exception as e:
            print(f"❌ Error obteniendo conocimiento aplicado: {e}")
            return None
    
    async def get_evaluation_feedback(self, evaluation_id: str) -> Optional[Dict[str, Any]]:
        """Obtener datos de feedback de una evaluación"""
        try:
            if not self.results_file.exists():
                return None
            
            wb = load_workbook(self.results_file, data_only=True)
            
            if RESULTS_SHEETS["feedback"]["name"] not in wb.sheetnames:
                return None
            
            ws = wb[RESULTS_SHEETS["feedback"]["name"]]
            
            # Buscar fila con el evaluation_id
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= len(FEEDBACK_COLUMNS) and row[0] == evaluation_id:
                    feedback_data = {}
                    for col_name, col_letter in FEEDBACK_COLUMNS.items():
                        col_index = self._get_col_index(col_letter)
                        if col_index < len(row):
                            feedback_data[col_name] = row[col_index]
                    wb.close()
                    return feedback_data
            
            wb.close()
            return None
            
        except Exception as e:
            print(f"❌ Error obteniendo feedback: {e}")
            return None
    
    # =================================================================
    # FUNCIONES AUXILIARES
    # =================================================================
    
    def _sanitize_evaluation_data(self, evaluation_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Sanitizar datos de evaluación para remover objetos AdminResponse embebidos
        y otros objetos problemáticos que puedan causar errores de validación
        """
        try:
            sanitized = {}
            
            for key, value in evaluation_data.items():
                # Saltar valores que son NaN o None
                if pd.isna(value) or value is None:
                    sanitized[key] = None
                    continue
                
                # Convertir a string y verificar que no sea un objeto serializado problemático
                str_value = str(value)
                
                # Detectar y filtrar objetos AdminResponse embebidos u otros objetos problemáticos
                if (
                    "AdminResponse" in str_value or
                    "default_factory" in str_value or
                    "timestamp" in str_value and "Field" in str_value or
                    str_value.startswith("{") and "success" in str_value and "message" in str_value
                ):
                    print(f"⚠️ Objeto problemático detectado en campo '{key}': {str_value[:100]}...")
                    # En lugar de saltar completamente, usar un valor por defecto
                    if key in ["evaluation_id", "cedula", "nombre", "procedure_codigo"]:
                        sanitized[key] = f"SANITIZED_{key.upper()}"
                    else:
                        sanitized[key] = None
                    continue
                
                # Limpiar valores con prefijos de enum
                if key == "campo" and "CampoEnum." in str_value:
                    # CampoEnum.cupiagua → cupiagua
                    clean_value = str_value.replace("CampoEnum.", "").capitalize()
                    sanitized[key] = clean_value
                elif key == "aprobo" and "SiNoEnum." in str_value:
                    # SiNoEnum.si → Sí, SiNoEnum.no → No
                    clean_value = str_value.replace("SiNoEnum.", "")
                    if clean_value == "si":
                        sanitized[key] = "Sí"
                    elif clean_value == "no":
                        sanitized[key] = "No"
                    else:
                        sanitized[key] = clean_value
                elif "OptionEnum." in str_value:
                    # OptionEnum.A → A
                    sanitized[key] = str_value.replace("OptionEnum.", "")
                else:
                    # Para tipos primitivos simples, mantener el valor
                    if isinstance(value, (str, int, float, bool)):
                        sanitized[key] = value
                    else:
                        # Para otros tipos, convertir a string de forma segura
                        sanitized[key] = str_value if str_value != "nan" else None
            
            return sanitized
            
        except Exception as e:
            print(f"❌ Error sanitizando datos de evaluación: {e}")
            print(f"❌ Datos problemáticos: {evaluation_data}")
            return None
    
    def _get_col_index(self, column_letter: str) -> int:
        """Convertir letra de columna a índice (A=0, B=1, etc.)"""
        return ord(column_letter.upper()) - ord('A')
    
    async def validate_data_file(self) -> Dict[str, Any]:
        """Validar archivo de datos y retornar información"""
        result = {
            "exists": False,
            "valid": False,
            "procedures_count": 0,
            "questions_count": 0,
            "errors": []
        }
        
        try:
            if not self.data_file.exists():
                result["errors"].append(f"Archivo no encontrado: {self.data_file}")
                return result
            
            result["exists"] = True
            
            # Verificar hojas
            excel_file = pd.ExcelFile(self.data_file)
            required_sheets = [DATA_SHEETS["procedures"]["name"], DATA_SHEETS["questions"]["name"]]
            
            for sheet in required_sheets:
                if sheet not in excel_file.sheet_names:
                    result["errors"].append(f"Hoja faltante: {sheet}")
            
            if result["errors"]:
                return result
            
            # Contar registros válidos
            procedures_df = pd.read_excel(self.data_file, sheet_name=DATA_SHEETS["procedures"]["name"])
            questions_df = pd.read_excel(self.data_file, sheet_name=DATA_SHEETS["questions"]["name"])
            
            # Contar solo filas con datos válidos
            valid_procedures = 0
            for _, row in procedures_df.iterrows():
                if not pd.isna(row.iloc[0]) and str(row.iloc[0]).strip() != "":
                    valid_procedures += 1
            
            valid_questions = 0
            for _, row in questions_df.iterrows():
                if not pd.isna(row.iloc[0]) and str(row.iloc[0]).strip() != "":
                    valid_questions += 1
            
            result["procedures_count"] = valid_procedures
            result["questions_count"] = valid_questions
            result["valid"] = valid_procedures > 0 and valid_questions > 0
            
        except Exception as e:
            result["errors"].append(f"Error validando archivo: {str(e)}")
        
        return result